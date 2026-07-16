#!/usr/bin/env python3
"""
Source candidates from a job description using the Exa Agent API — end to end.

Pipeline: graded discovery across talent segments  ->  high-effort verification  ->
calibrated ranking  ->  candidates.csv + candidates.html (interactive viewer;
+ candidates.xlsx if openpyxl is installed).

Self-contained (Python stdlib only; openpyxl optional for the .xlsx).

USAGE
  export EXA_API_KEY=...                      # or run scripts/set-exa-key.sh once (stores ~/.config/exa/key)
  python source_candidates.py --config config.json
  python source_candidates.py --config config.json --no-verify --limit-segments 1   # quick smoke run
  python source_candidates.py --config config.json --more    # continue: fetch more, dedupe, keep verdicts
  python source_candidates.py --config config.json --exclude-file pipeline.txt   # skip an existing list

Edit config.json to point at any role (see config.example.json). The driving agent
(Claude Code / Devin) fills the config from the JD: rubric dimensions, segments,
locations, and an optional exclude_employer.
"""
import os, sys, json, time, threading, urllib.request, urllib.error, csv, re, argparse

BASE = "https://api.exa.ai"


def _resolve_key():
    """Env var first, then the credentials file written by scripts/set-exa-key.sh.
    Lets non-technical users run one setup script instead of editing a shell profile."""
    k = os.environ.get("EXA_API_KEY")
    if k and k.strip():
        return k.strip()
    path = os.path.expanduser(os.environ.get("EXA_KEY_FILE", "~/.config/exa/key"))
    try:
        with open(path) as f:
            k = f.read().strip()
            return k or None
    except OSError:
        return None


KEY = _resolve_key()
CAP = ["strong", "partial", "none", "unknown"]          # capability scale
STR = ["none", "weak", "medium", "strong", "unknown"]   # signal-strength scale
ARR = {"type": "array", "items": {"type": "string"}}


# ----------------------------- HTTP -----------------------------
def _req(method, path, body=None):
    if not KEY:
        sys.exit("ERROR: no Exa API key found. Set EXA_API_KEY in your environment, "
                 "or run scripts/set-exa-key.sh (set-exa-key.ps1 on Windows) to store one "
                 "in ~/.config/exa/key.")
    data = json.dumps(body).encode() if body is not None else None
    r = urllib.request.Request(BASE + path, data=data,
                               headers={"x-api-key": KEY, "Content-Type": "application/json"}, method=method)
    try:
        with urllib.request.urlopen(r, timeout=120) as resp:
            return resp.status, json.loads(resp.read().decode())
    except urllib.error.HTTPError as e:
        return e.code, {"error": e.read().decode()}


def create_run(body, label=""):
    """Start a run; one short retry on 429/5xx so a transient rejection does not drop a segment."""
    st, resp = _req("POST", "/agent/runs", body)
    if st == 429 or st >= 500:
        print(f"  [{label}] create hit HTTP {st}, retrying once...")
        time.sleep(2)
        st, resp = _req("POST", "/agent/runs", body)
    if st >= 300:
        print(f"  [{label}] create failed HTTP {st}: {str(resp)[:160]}")
        return None
    return resp.get("id")


def wait_run(rid, max_wait=1200, poll=8, label=""):
    """Poll a run to terminal; cancel + return None on hang/failure."""
    t0 = time.time()
    while True:
        st, r = _req("GET", "/agent/runs/" + rid)
        if 400 <= st < 500 and st != 429:   # bad/expired run id: not worth polling for 20 min
            print(f"  [{label}] poll failed HTTP {st}: {str(r)[:120]}")
            return None
        status = r.get("status")
        if status == "completed":
            return r
        if status in ("failed", "canceled", "cancelled"):
            print(f"  [{label}] run {status} (stopReason={r.get('stopReason')})")
            return None
        if time.time() - t0 > max_wait:
            _req("POST", f"/agent/runs/{rid}/cancel")   # stop the stuck run
            print(f"  [{label}] timed out after {max_wait}s, cancelled {rid}")
            return None
        time.sleep(poll)


# ----------------------------- schema ---------------------------
def _grade(scale, extra=None):
    props = {"level": {"type": "string", "enum": scale}, "signals": ARR}
    req = ["level", "signals"]
    for e in (extra or []):
        props[e] = ARR
        req.append(e)
    return {"type": "object", "additionalProperties": False, "required": req, "properties": props}


CONTACT_FORMATS = {"email": "email", "phone": "phone", "uri": "uri", "url": "uri"}


def contact_fields(cfg):
    """Normalize opt-in contact fields into (column name, JSON Schema format) pairs."""
    out = []
    for field in cfg.get("contact_fields", []):
        if isinstance(field, str):
            key = field.strip()
            fmt = CONTACT_FORMATS.get(key.lower())
            if not fmt:
                raise ValueError(f"unknown contact field {field!r}; use email/phone/uri or "
                                 "{'key': ..., 'format': ...}")
        elif isinstance(field, dict):
            key, fmt = field["key"], field.get("format", "email")
        else:
            continue
        out.append((key, fmt))
    return out


def build_schema(cfg):
    props = {
        "name": {"type": "string"}, "currentTitle": {"type": "string"}, "currentCompany": {"type": "string"},
        "location": {"type": ["string", "null"]}, "linkedinUrl": {"type": ["string", "null"]},
        "yearsRelevantExperience": {"type": ["number", "null"]},
    }
    req = list(props)
    for key, fmt in contact_fields(cfg):
        props[key] = {"type": ["string", "null"], "format": fmt}
        req.append(key)
    if cfg.get("exclude_employer"):
        props["currentlyAtExcludedEmployer"] = {"type": "boolean"}
        req.append("currentlyAtExcludedEmployer")
    for d in cfg["dimensions"]:
        scale = CAP if d.get("scale", "capability") == "capability" else STR
        props[d["key"]] = _grade(scale, d.get("extra"))
        req.append(d["key"])
    props["seniority"] = {"type": "object", "additionalProperties": False, "required": ["level", "signals"],
                          "properties": {"level": {"type": "string", "enum": ["ic_mid", "ic_senior",
                          "ic_staff_principal", "manager", "director_plus", "unknown"]}, "signals": ARR}}
    props["overallFit"] = {"type": "object", "additionalProperties": False,
                           "required": ["tier", "confidence", "signalsUsed", "concerns"],
                           "properties": {"tier": {"type": "string", "enum": ["exceptional", "strong",
                           "moderate", "weak", "unknown"]}, "confidence": {"type": "string", "enum": ["low",
                           "medium", "high"]}, "signalsUsed": ARR, "concerns": ARR}}
    props["mobility"] = {"type": "object", "additionalProperties": False,
                         "required": ["monthsInCurrentRole", "monthsAtCurrentCompany",
                                      "avgMonthsPerPriorRole", "seniorityVsRole", "signals"],
                         "properties": {"monthsInCurrentRole": {"type": ["number", "null"]},
                                        "monthsAtCurrentCompany": {"type": ["number", "null"]},
                                        "avgMonthsPerPriorRole": {"type": ["number", "null"]},
                                        "seniorityVsRole": {"type": "string", "enum": ["step_up",
                                        "aligned", "step_down", "unknown"]}, "signals": ARR}}
    req += ["seniority", "overallFit", "mobility"]
    return {"$schema": "https://json-schema.org/draft/2020-12/schema", "type": "object",
            "additionalProperties": False, "required": ["candidates"],
            "properties": {"candidates": {"type": "array", "maxItems": cfg.get("max_per_call", 12),
                           "items": {"type": "object", "additionalProperties": False,
                                     "required": req, "properties": props}}}}


# ----------------------------- queries --------------------------
def _providers(cfg):
    return [d.strip() for d in cfg.get("data_sources", []) if isinstance(d, str) and d.strip()]


def discovery_query(cfg, segment):
    p = [f"Find real candidates for: {cfg['role']}.",
         "MUST-HAVE profile: " + " ".join(cfg.get("rubric_must_haves", [])),
         "NICE-TO-HAVE signals: " + " ".join(cfg.get("rubric_signals", [])),
         f"FOCUS for this search (verify independently — do NOT treat as ground truth): {segment['focus']}",
         f"Strongly prioritize people based in or near: {', '.join(cfg.get('locations', [])) or 'any location'}."]
    if cfg.get("exclude_employer"):
        p.append(f"EXCLUDE anyone currently employed at {cfg['exclude_employer']}; set "
                 "currentlyAtExcludedEmployer=true if they are and then do NOT include them. Do not bias toward "
                 "that employer's own job-title vocabulary — value the transferable profile and equivalent roles "
                 "at other companies.")
    p.append("For EACH candidate, grade every rubric dimension with a level and the concrete public signals that "
             "justify it, plus an overallFit (tier, confidence, signalsUsed, concerns).")
    p.append("For EACH candidate, also fill mobility from their dated public work history: monthsInCurrentRole "
             "(months since they started the current position), monthsAtCurrentCompany, avgMonthsPerPriorRole "
             "(mean months per position across roughly their last 3-5 previous positions), and seniorityVsRole "
             "(would this role be a step_up, aligned, or step_down versus their current level; step_down means "
             "they are overqualified). Put the dated evidence in mobility.signals. Use null or 'unknown' when "
             "start dates are not public; never estimate a tenure without a dated source.")
    p.append("Grade strictly and comparatively so the ranking discriminates: a dimension is 'strong' only with "
             "direct public evidence, 'partial' when the evidence is indirect or inferred. Reserve tier "
             "'exceptional' for at most one or two near-perfect fits per batch; a good candidate is tier "
             "'strong' with a mix of strong and partial grades. Set confidence 'high' only when multiple "
             "independent sources corroborate.")
    if _providers(cfg):
        p.append(f"A {' and '.join(_providers(cfg))} people data source is attached; use it to surface "
                 "candidates and to confirm each person's current title, employer, location, and LinkedIn URL.")
    contacts = contact_fields(cfg)
    if contacts:
        p.append("Also return these public contact fields for each candidate: "
                 + ", ".join(key for key, _ in contacts)
                 + ". Set a field to null when no value can be confirmed; never guess or fabricate a contact value.")
    p.append("Search beyond the LinkedIn headline: full work history, GitHub, conference/meetup talks, company "
             "team pages, certification registries, blogs; corroborate across at least two sources where possible.")
    p.append("Use null, empty arrays, or the 'unknown' enum whenever a fact is not supported by public evidence; "
             "NEVER fabricate a name, LinkedIn URL, employer, or number. If you cannot confirm a real LinkedIn "
             "profile, set linkedinUrl to null.")
    p.append(f"Return up to {cfg.get('max_per_call', 12)} real candidates.")
    return "\n\n".join(p)


def verify_query(cfg):
    head = (f"Fact-check this recruiting shortlist for: {cfg['role']}. The input data has one row per person "
            "with an id and their claimed name, title, company, location, and LinkedIn URL. For EACH row, use "
            "web search to determine: (1) are they a real, currently-active professional matching the claimed "
            "name/title/company; (2) does the LinkedIn URL plausibly belong to them; (3) how well does their real "
            "background match the role")
    if cfg.get("exclude_employer"):
        head += f"; (4) do they CURRENTLY work at {cfg['exclude_employer']} (set currently_excluded=true if so)"
    head += (". Be skeptical: if you cannot find evidence, mark exists 'uncertain' or 'not_found'. Return "
             "exactly one verdict per row, copying the row's id into the verdict's id field unchanged.")
    return head


def verify_schema(cfg):
    """Verdicts echo each input row's id, so results join onto exactly the right candidate
    even when two people share a name. currently_excluded is only offered when an employer
    exclusion was requested; otherwise agents repurpose it for unrelated doubts."""
    props = {"id": {"type": "string"}, "name": {"type": "string"},
             "exists": {"type": "string", "enum": ["confirmed", "likely", "uncertain", "not_found"]},
             "linkedin_valid": {"type": "string", "enum": ["valid", "unverifiable", "wrong"]},
             "matches_role": {"type": "string", "enum": ["strong", "partial", "weak", "no"]},
             "verified_title_company": {"type": "string"}}
    if cfg.get("exclude_employer"):
        props["currently_excluded"] = {"type": "boolean"}
    return {"$schema": "https://json-schema.org/draft/2020-12/schema", "type": "object",
            "additionalProperties": False, "required": ["verdicts"],
            "properties": {"verdicts": {"type": "array", "items": {
                "type": "object", "additionalProperties": False, "required": ["id", "name", "exists", "matches_role"],
                "properties": props}}}}


# ----------------------------- orchestration --------------------
def _norm_li(u):
    u = (u or "").lower().split("?")[0].rstrip("/")
    m = re.search(r"linkedin\.com/in/([^/]+)", u)
    return "li:" + m.group(1) if m else ""


def _key(c):
    return _norm_li(c.get("linkedinUrl")) or ("nm:" + re.sub(r"[^a-z ]", "", (c.get("name") or "").lower()).strip())


def discovery_body(cfg, schema, seg, excl, prev_rid=None):
    if prev_rid:
        q = (f"Find up to {cfg.get('max_per_call', 12)} MORE candidates matching the same brief, excluding "
             "everyone already returned and everyone in the exclusion list. Apply the same rubric grading, "
             "mobility (dated work-history) fields, and anti-fabrication rules as before.")
        if cfg.get("exclude_employer"):
            q += f" Still exclude anyone currently employed at {cfg['exclude_employer']}."
    else:
        q = discovery_query(cfg, seg)
    body = {"query": q, "effort": cfg.get("discovery_effort", "medium"), "outputSchema": schema}
    if prev_rid:
        body["previousRunId"] = prev_rid
    if excl:
        body["input"] = {"exclusion": excl}
    if _providers(cfg):
        body["dataSources"] = [{"provider": p} for p in _providers(cfg)]
    return body


def discover(cfg, schema, segments, concurrency, prev_run_ids=None, seen_names=None):
    """One run per segment. Returns (candidates, {segment label: run id}); the run ids
    feed --more continuations via previousRunId."""
    found, seen_names = [], list(seen_names or [])
    run_ids = {}
    for i in range(0, len(segments), concurrency):
        batch = segments[i:i + concurrency]
        # newest 300: keeps continuation-round create bodies bounded
        excl = [{"person": n} for n in seen_names[-300:]]
        # Create sequentially (a parallel create burst can trip the account QPS limit),
        # then poll the started runs concurrently.
        started = []
        for seg in batch:
            rid = create_run(discovery_body(cfg, schema, seg, excl, (prev_run_ids or {}).get(seg["label"])),
                             label=seg["label"])
            if rid:
                started.append((seg, rid))
                run_ids[seg["label"]] = rid
            time.sleep(1)
        results = {}

        def work(seg, rid):
            r = wait_run(rid, label=seg["label"])
            out = ((r or {}).get("output") or {})
            cands = (out.get("structured") or {}).get("candidates") or []
            grounded = grounding_by_index(out.get("grounding"))
            for i, c in enumerate(cands):
                c["_segment"] = seg["label"]
                if i in grounded:
                    c["_sources"] = grounded[i]
            results[seg["label"]] = cands
            print(f"  [{seg['label']}] {len(cands)} candidates")

        ts = [threading.Thread(target=work, args=(s, rid)) for s, rid in started]
        for t in ts: t.start()
        for t in ts: t.join()
        for s, _rid in started:
            for c in results.get(s["label"], []):
                found.append(c)
                if c.get("name"):
                    seen_names.append(c["name"])
    return found, run_ids


_CAND_FIELD = re.compile(r"^structured\.candidates\[(\d+)\](?:\.|$)")


def grounding_by_index(grounding):
    """Attribute a run's output.grounding citations to candidate indexes. Field paths arrive at
    three granularities: "structured.candidates[3]" (whole candidate), "structured.candidates[3].name"
    (one field, rolled up to its candidate), and bare "structured" (run-level research trail,
    dropped: it supports no one candidate's claims). Citation URLs dedup per candidate.
    Returns {index: urls}."""
    by = {}
    if not isinstance(grounding, list):
        return by
    for e in grounding:
        if not isinstance(e, dict):
            continue
        m = _CAND_FIELD.match(e.get("field") or "")
        cites = e.get("citations")
        if not m or not isinstance(cites, list):
            continue
        i = int(m.group(1))
        urls = by.setdefault(i, [])
        for c in cites:
            u = c.get("url") if isinstance(c, dict) else None
            if isinstance(u, str) and u.strip() and u not in urls:
                urls.append(u)
    return {i: urls for i, urls in by.items() if urls}


def excluded_people(cfg, path=None):
    """People the recruiter wants excluded — an existing pipeline/ATS list. Merges the
    `exclude_people` config key with an optional --exclude-file (one entry per line).
    Each entry is a plain string (a person's name OR a LinkedIn URL) or a
    {name, linkedinUrl} object. Returns (names, keys): `names` seed every discovery run's
    input.exclusion so the search avoids them up front; `keys` (normalized the same way as
    _key) drop any that still come back."""
    entries = list(cfg.get("exclude_people") or [])
    if path:
        with open(path) as f:
            entries += [ln.strip() for ln in f if ln.strip()]
    names, keys = [], set()
    for e in entries:
        name = url = None
        if isinstance(e, dict):
            name, url = e.get("name"), e.get("linkedinUrl")
        elif isinstance(e, str):
            if "linkedin.com/in/" in e.lower():
                url = e
            else:
                name = e
        if url:
            k = _norm_li(url)
            if k:
                keys.add(k)
        if name:
            names.append(name)
            keys.add("nm:" + re.sub(r"[^a-z ]", "", name.lower()).strip())
    return names, keys


def is_excluded(cfg, c):
    if not cfg.get("exclude_employer"):
        return False
    if c.get("currentlyAtExcludedEmployer") is True:
        return True
    terms = [t.strip().lower() for t in re.split(r"[\/,]", cfg["exclude_employer"]) if t.strip()]
    co = (c.get("currentCompany") or "").lower()
    return any(t in co for t in terms)


def lvl(c, k):
    return ((c.get(k) or {}).get("level")) or "unknown"


def score(cfg, c):
    capw = {"strong": 2, "partial": 1, "none": 0, "unknown": 0}
    strw = {"strong": 2, "medium": 1.5, "weak": 1, "none": 0, "unknown": 0}
    s = 0
    for d in cfg["dimensions"]:
        w = capw if d.get("scale", "capability") == "capability" else strw
        s += w.get(lvl(c, d["key"]), 0)
    tier = {"exceptional": 90, "strong": 76, "moderate": 58, "weak": 38, "unknown": 50}.get(
        (c.get("overallFit") or {}).get("tier", "unknown"), 50)
    conf = {"high": 6, "medium": 0, "low": -6}.get((c.get("overallFit") or {}).get("confidence", "medium"), 0)
    loc = (c.get("location") or "").lower()
    metro = any(w.lower().split(",")[0][:6] in loc for w in cfg.get("locations", [])) if loc else False
    return tier + conf + s * 1.4 + (4 if metro else 0), metro


def likely_to_move(c):
    """0-100 propensity to switch jobs from the mobility signals; display-only
    (never blended into the match score or ranking). None when no signal is known,
    so the output shows a blank instead of a fake neutral 50."""
    m = c.get("mobility") or {}
    def num(v):
        return v if isinstance(v, (int, float)) and v >= 0 else None
    tenure = num(m.get("monthsInCurrentRole"))
    if tenure is None:
        tenure = num(m.get("monthsAtCurrentCompany"))
    cadence = num(m.get("avgMonthsPerPriorRole"))
    vs_role = m.get("seniorityVsRole") if m.get("seniorityVsRole") in ("step_up", "aligned", "step_down") else "unknown"
    if tenure is None and cadence is None and vs_role == "unknown":
        return None
    s = 50
    if tenure is not None:
        if tenure < 9: s -= 25
        elif tenure < 18: s -= 10
        elif tenure <= 42: s += 20
        elif tenure <= 72: s += 8
        else: s -= 8
    if cadence is not None:
        if cadence <= 30: s += 15
        elif cadence <= 48: s += 5
        else: s -= 8
    s += {"step_down": -20, "step_up": 10, "aligned": 5}.get(vs_role, 0)
    return round(min(100, max(0, s)))


def max_possible_score(cfg):
    """Displayed scores are a percentage of the best score the rubric allows, not a clip at
    100: clipping collapsed every good candidate onto the same number, while a percentage
    keeps the tier gaps visible."""
    return 90 + 6 + len(cfg["dimensions"]) * 2 * 1.4 + (4 if cfg.get("locations") else 0)


def calibrate(cfg, c):
    """Discount unconfirmed identity/location/low confidence so thin profiles don't top the list."""
    base, metro = c["_score"], c["_metro"]
    co = (c.get("currentCompany") or "").strip().lower()
    pen = 0
    if co in ("", "unknown", "n/a"): pen += 15
    if not (c.get("location") or "").strip(): pen += 10
    elif not metro and cfg.get("locations"): pen += 6
    # low confidence is discounted beyond its base adjustment; high is already
    # rewarded in the base score, so no double count here
    if (c.get("overallFit") or {}).get("confidence") == "low": pen += 10
    if c.get("_exists") == "not_found": pen += 25
    elif c.get("_exists") == "uncertain": pen += 8
    cal = (base - pen) * 100 / max_possible_score(cfg)
    if co in ("", "unknown", "n/a") or not (c.get("location") or "").strip():
        cal = min(cal, 82)
    return round(min(100, max(0, cal)))


def verify(cfg, candidates, concurrency):
    """Fact-check in batches of 8. Each shortlist row travels as an input.data row carrying
    the candidate's stable key as id; the verdict echoes it back, so the join is exact even
    when two people share a name. Returns {candidate key: verdict}."""
    batches = [candidates[i:i + 8] for i in range(0, len(candidates), 8)]
    schema = verify_schema(cfg)
    verdicts, lock = {}, threading.Lock()
    for i in range(0, len(batches), concurrency):
        grp = batches[i:i + concurrency]
        started = []
        for b in grp:
            rows = [{"id": _key(c), "name": c.get("name"), "claimed_title": c.get("currentTitle"),
                     "claimed_company": c.get("currentCompany"), "claimed_location": c.get("location"),
                     "linkedin_url": c.get("linkedinUrl")} for c in b]
            rid = create_run({"query": verify_query(cfg), "effort": cfg.get("verify_effort", "high"),
                              "outputSchema": schema, "input": {"data": rows}}, label="verify")
            if rid:
                started.append((b, rid))
            time.sleep(1)

        def work(b, rid):
            r = wait_run(rid, label="verify")
            sent = {_key(c) for c in b}
            for v in (((r or {}).get("output") or {}).get("structured") or {}).get("verdicts") or []:
                if v.get("id") in sent:
                    with lock:
                        verdicts[v["id"]] = v
                else:
                    print(f"  dropped verdict with unknown id {v.get('id')!r} ({v.get('name')})")

        ts = [threading.Thread(target=work, args=(b, rid)) for b, rid in started]
        for t in ts: t.start()
        for t in ts: t.join()
        print(f"  verified {min((i + concurrency) * 8, len(candidates))}/{len(candidates)}")
    return verdicts


# ----------------------------- output ---------------------------
def write_outputs(cfg, final):
    dims = [d["key"] for d in cfg["dimensions"]]
    contacts = [key for key, _ in contact_fields(cfg)]
    cols = ["rank", "name", "linkedinUrl", "currentTitle", "currentCompany", "location",
            "score", "likely_to_move",
            "months_in_current_role", "avg_months_per_prior_role", "seniority_vs_role",
            "mobility_signals", "overall_tier", "confidence"] \
        + dims + contacts + ["seniority", "concerns", "verify_exists", "verify_match",
                  "sources", "segment"]

    def row(i, c):
        of = c.get("overallFit") or {}
        ltm = likely_to_move(c)
        mob = c.get("mobility") or {}
        months = mob.get("monthsInCurrentRole")
        cadence = mob.get("avgMonthsPerPriorRole")
        vs_role = mob.get("seniorityVsRole")
        return [i, c.get("name"), c.get("linkedinUrl"), c.get("currentTitle"), c.get("currentCompany"),
                c.get("location"),
                c["_calib"], "" if ltm is None else ltm,
                months if isinstance(months, (int, float)) else "",
                cadence if isinstance(cadence, (int, float)) else "",
                vs_role if vs_role in ("step_up", "aligned", "step_down") else "",
                " | ".join((mob.get("signals") or [])[:3]),
                of.get("tier"), of.get("confidence")] + [lvl(c, k) for k in dims] \
            + [c.get(key) for key in contacts] \
            + [(c.get("seniority") or {}).get("level"), " | ".join((of.get("concerns") or [])[:2]),
               c.get("_exists", ""), c.get("_match", ""),
               " | ".join(c.get("_sources") or []), c.get("_segment")]

    with open("candidates.csv", "w", newline="") as fh:
        w = csv.writer(fh); w.writerow(cols)
        for i, c in enumerate(final, 1):
            w.writerow(row(i, c))
    print(f"wrote candidates.csv ({len(final)} rows)")
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Candidates"
        ws.append(cols)
        for c in ws[1]:
            c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor="1F3864")
        for i, c in enumerate(final, 1):
            ws.append(row(i, c))
        ws.freeze_panes = "C2"
        wb.save("candidates.xlsx"); print("wrote candidates.xlsx")
    except ImportError:
        print("(openpyxl not installed — skipped .xlsx; `pip install openpyxl` for the formatted sheet)")
    try:
        try:
            import render_viewer
        except ImportError:
            sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
            import render_viewer
        render_viewer.render("candidates.csv", "candidates.html", title=cfg.get("role") or "Candidates")
    except (Exception, SystemExit) as e:
        print(f"(candidates.html skipped: {e})")


# ----------------------------- main -----------------------------
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--config", required=True)
    ap.add_argument("--target", type=int, default=50, help="how many candidates to keep")
    ap.add_argument("--no-verify", action="store_true")
    ap.add_argument("--limit-segments", type=int, default=0, help="use only the first N segments (smoke test)")
    ap.add_argument("--max-concurrent", type=int, default=0, help="override config concurrency")
    ap.add_argument("--more", action="store_true",
                    help="continue the previous session's runs (previousRunId) and fetch new people")
    ap.add_argument("--state", default="sourcing_state.json", help="session state file, enables --more")
    ap.add_argument("--exclude-file",
                    help="file of names or LinkedIn URLs (one per line) to exclude — an existing "
                         "pipeline/ATS list; merged with the config's exclude_people")
    a = ap.parse_args()
    cfg = json.load(open(a.config))
    segments = cfg["segments"][:a.limit_segments] if a.limit_segments else cfg["segments"]
    conc = a.max_concurrent or cfg.get("concurrency", 2)
    schema = build_schema(cfg)

    state = {"run_ids": {}, "pool": [], "verdicts": {}}
    if a.more:
        if not os.path.exists(a.state):
            sys.exit(f"ERROR: --more continues a previous session, but {a.state} was not found.")
        state = json.load(open(a.state))

    excl_names, excl_keys = excluded_people(cfg, a.exclude_file)
    if excl_keys:
        print(f"== excluding {len(excl_keys)} people from the recruiter's list ==")

    by = {_key(c): c for c in state["pool"] if _key(c) not in excl_keys}
    # excl_names lead so they seed input.exclusion even on the first (non-continuation) run
    seen = excl_names + [c["name"] for c in by.values() if c.get("name")]

    print(f"== discovery: {len(segments)} segments, {conc} concurrent{' (continuation)' if a.more else ''} ==")
    raw, run_ids = discover(cfg, schema, segments, conc,
                            prev_run_ids=state["run_ids"] if a.more else None, seen_names=seen)

    # dedup + drop excluded + score
    for c in raw:
        if is_excluded(cfg, c) or _key(c) in excl_keys:
            continue
        c["_score"], c["_metro"] = score(cfg, c)
        k = _key(c)
        if k.replace("nm:", "").strip() and (k not in by or c["_score"] > by[k]["_score"]):
            by[k] = c
    pool = sorted(by.values(), key=lambda x: -x["_score"])
    print(f"== {len(raw)} raw -> {len(pool)} unique (non-excluded) ==")

    verdicts = state.get("verdicts", {})
    shortlist = pool[:max(a.target + 14, a.target)]
    if not a.no_verify and shortlist:
        unchecked = [c for c in shortlist if _key(c) not in verdicts]
        if unchecked:
            print(f"== verification: {len(unchecked)} of top {len(shortlist)} (effort {cfg.get('verify_effort','high')}) ==")
            verdicts.update(verify(cfg, unchecked, conc))
    for c in shortlist:
        v = verdicts.get(_key(c), {})
        c["_exists"], c["_match"] = v.get("exists", "unchecked"), v.get("matches_role", "unchecked")
        c["_vexcl"] = v.get("currently_excluded", None)
    erank = {"confirmed": 0, "likely": 1, "uncertain": 2, "unchecked": 2, "not_found": 3}
    mrank = {"strong": 0, "partial": 1, "weak": 2, "unchecked": 1, "no": 3}

    def eligible(c):
        return (c.get("_exists", "unchecked") != "not_found" and c.get("_match", "unchecked") != "no"
                and c.get("_vexcl") is not True
                and not (not (c.get("linkedinUrl") or "").strip()
                         and (c.get("currentCompany") or "").strip().lower() in ("", "unknown", "n/a")))

    for c in shortlist:
        c["_calib"] = calibrate(cfg, c)
    elig = sorted([c for c in shortlist if eligible(c)],
                  key=lambda c: (erank.get(c.get("_exists", "unchecked"), 2),
                                 mrank.get(c.get("_match", "unchecked"), 1), -c["_calib"]))
    final = elig[:a.target]
    print(f"== final: {len(final)} candidates ==")
    write_outputs(cfg, final)

    # newest run ids win so the next --more continues from the latest round
    json.dump({"run_ids": {**state.get("run_ids", {}), **run_ids}, "pool": pool, "verdicts": verdicts},
              open(a.state, "w"))
    print(f"wrote {a.state} (rerun with --more to fetch additional candidates)")


if __name__ == "__main__":
    main()
