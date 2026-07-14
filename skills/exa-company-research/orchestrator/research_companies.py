#!/usr/bin/env python3
"""
Build a ranked, verified company list from a research plan using the Exa Agent API — end to end.

Pipeline: discovery across segment fan-out (hard criteria filter, soft criteria grade,
freeform data columns, optional Exa Connect providers)  ->  dedupe (by domain, and against an
existing list)  ->  high-effort verification of every hard criterion  ->  calibrated ranking
->  companies.csv + companies.html (interactive viewer; + companies.xlsx if openpyxl is
installed).

Self-contained (Python stdlib only; openpyxl optional for the .xlsx).

USAGE
  export EXA_API_KEY=...                        # or run scripts/set-exa-key.sh once (stores ~/.config/exa/key)
  python research_companies.py --config config.json
  python research_companies.py --config config.json --no-verify --limit-segments 1  # smoke run
  python research_companies.py --config config.json --more   # continue: fetch more, dedupe, keep verdicts
  python research_companies.py --config config.json --rounds 3   # exhaustive lists: keep digging until dry
  python research_companies.py --config config.json --exclude-file known.txt  # skip an existing list

Edit config.json per ask (see config.example.json). The driving agent (Claude Code / Devin)
fills the config from the user's ask: hard criteria, soft criteria, columns, providers,
segments, existing list.
"""
import os, ssl, sys, json, time, threading, urllib.request, urllib.error, csv, re, argparse

BASE = "https://api.exa.ai"


def _resolve_key():
    """Env var first, then the credentials file written by scripts/set-exa-key.sh.
    Lets non-technical users run one setup script instead of editing a shell profile."""
    k = os.environ.get("EXA_API_KEY")
    if k and k.strip():
        return k.strip()
    path = os.path.expanduser(os.environ.get("EXA_KEY_FILE", "~/.config/exa/key"))
    try:
        with open(path) as f:
            k = f.read().strip()
            return k or None
    except OSError:
        return None


KEY = _resolve_key()
CAP = ["strong", "partial", "none", "unknown"]          # capability scale
STR = ["none", "weak", "medium", "strong", "unknown"]   # signal-strength scale
ARR = {"type": "array", "items": {"type": "string"}}
LEGAL_SUFFIXES = ("inc", "llc", "ltd", "limited", "corp", "corporation", "co",
                  "gmbh", "plc", "sas", "bv", "ab", "oy", "pty", "srl", "sa")


# ----------------------------- HTTP -----------------------------
_SSL_CTX = None


def _ssl_context():
    """Default context, with certifi's roots when available — python.org macOS builds
    often ship without local root certs and fail every HTTPS call otherwise."""
    global _SSL_CTX
    if _SSL_CTX is None:
        _SSL_CTX = ssl.create_default_context()
        try:
            import certifi
            _SSL_CTX.load_verify_locations(certifi.where())
        except ImportError:
            pass
    return _SSL_CTX


def _req(method, path, body=None):
    if not KEY:
        sys.exit("ERROR: no Exa API key found. Set EXA_API_KEY in your environment, "
                 "or run scripts/set-exa-key.sh (set-exa-key.ps1 on Windows) to store one "
                 "in ~/.config/exa/key.")
    data = json.dumps(body).encode() if body is not None else None
    r = urllib.request.Request(BASE + path, data=data,
                               headers={"x-api-key": KEY, "Content-Type": "application/json"}, method=method)
    try:
        with urllib.request.urlopen(r, timeout=120, context=_ssl_context()) as resp:
            return resp.status, json.loads(resp.read().decode())
    except urllib.error.HTTPError as e:
        return e.code, {"error": e.read().decode()}
    except (urllib.error.URLError, TimeoutError, OSError) as e:
        if isinstance(getattr(e, "reason", None), ssl.SSLCertVerificationError):
            sys.exit("ERROR: HTTPS certificate verification failed. Your Python install is "
                     "missing root certificates — `pip install certifi`, or on macOS run "
                     "'Install Certificates.command' from the Python application folder.")
        return 0, {"error": str(e)}   # transient network error: create_run retries once


def create_run(body, label=""):
    """Start a run; one short retry on 429/5xx so a transient rejection does not drop a segment."""
    st, resp = _req("POST", "/agent/runs", body)
    if st in (0, 429) or st >= 500:
        print(f"  [{label}] create hit {'network error' if st == 0 else 'HTTP %d' % st}, retrying once...")
        time.sleep(2)
        st, resp = _req("POST", "/agent/runs", body)
    if st == 0 or st >= 300:
        print(f"  [{label}] create failed ({st or 'network error'}): {str(resp)[:160]}")
        return None
    return resp.get("id")


def wait_run(rid, max_wait=1200, poll=8, label=""):
    """Poll a run to terminal; cancel + return None on hang/failure."""
    t0 = time.time()
    while True:
        st, r = _req("GET", "/agent/runs/" + rid)
        if 400 <= st < 500 and st != 429:   # bad/expired run id: not worth polling for 20 min
            print(f"  [{label}] poll failed HTTP {st}: {str(r)[:120]}")
            return None
        status = r.get("status")
        if status == "completed":
            return r
        if status in ("failed", "canceled", "cancelled"):
            print(f"  [{label}] run {status} (stopReason={r.get('stopReason')})")
            return None
        if time.time() - t0 > max_wait:
            _req("POST", f"/agent/runs/{rid}/cancel")   # stop the stuck run
            print(f"  [{label}] timed out after {max_wait}s, cancelled {rid}")
            return None
        time.sleep(poll)


# ----------------------------- config validation ----------------
# Field names the pipeline itself owns; a plan key colliding with one would clobber the
# schema or produce duplicate CSV headers the viewer silently misreads.
RESERVED_KEYS = {"name", "website", "hq", "description", "overallFit", "rank", "company",
                 "score", "overall_tier", "confidence", "concerns", "verify_exists",
                 "verify_website", "corrections", "sources", "segment"}


def validate_config(cfg):
    problems = []
    keys = [k["key"] for k in cfg.get("hard_criteria", []) + cfg.get("soft_criteria", [])
            + cfg.get("columns", [])]
    for k in keys:
        if k in RESERVED_KEYS:
            problems.append(f"key {k!r} is reserved by the pipeline — rename it")
    dupes = {k for k in keys if keys.count(k) > 1}
    if dupes:
        problems.append(f"duplicate criterion/column keys: {', '.join(sorted(dupes))}")
    labels = [s["label"] for s in cfg.get("segments", [])]
    if len(labels) != len(set(labels)):
        problems.append("segment labels must be unique (results and continuations key on them)")
    if not cfg.get("hard_criteria"):
        problems.append("at least one hard criterion is required")
    if problems:
        sys.exit("ERROR in config:\n  - " + "\n  - ".join(problems))


# ----------------------------- schema ---------------------------
def _col_schema(col):
    """A plan column -> a nullable typed JSON-schema field whose description names its source
    (that description is how the agent routes the field to an attached Exa Connect provider)."""
    t = (col.get("type") or "string").replace("|null", "").strip()
    desc = col.get("desc", "")
    if col.get("source") and col["source"] != "web":
        desc = (desc + f" (from {col['source']})").strip()
    if t == "string[]":
        return {"type": ["array", "null"], "items": {"type": "string"}, "description": desc}
    if t not in ("string", "number", "boolean"):
        t = "string"
    return {"type": [t, "null"], "description": desc}


def build_schema(cfg):
    props = {
        "name": {"type": "string", "description": "official company name"},
        "website": {"type": ["string", "null"], "description": "canonical homepage URL, null if unconfirmed"},
        "hq": {"type": ["string", "null"], "description": "HQ city/region"},
        "description": {"type": "string", "description": "one line: what the company actually sells"},
    }
    req = list(props)
    for h in cfg["hard_criteria"]:
        props[h["key"]] = {"type": "object", "additionalProperties": False,
                           "required": ["met", "evidence"], "description": h["text"],
                           "properties": {"met": {"type": "string", "enum": ["yes", "no", "unknown"]},
                                          "evidence": ARR}}
        req.append(h["key"])
    for s in cfg.get("soft_criteria", []):
        scale = CAP if s.get("scale", "strength") == "capability" else STR
        props[s["key"]] = {"type": "object", "additionalProperties": False,
                           "required": ["level", "signals"], "description": s["text"],
                           "properties": {"level": {"type": "string", "enum": scale}, "signals": ARR}}
        req.append(s["key"])
    for c in cfg.get("columns", []):
        props[c["key"]] = _col_schema(c)
        req.append(c["key"])
    props["overallFit"] = {"type": "object", "additionalProperties": False,
                           "required": ["tier", "confidence", "signalsUsed", "concerns"],
                           "properties": {"tier": {"type": "string", "enum": ["exceptional", "strong",
                           "moderate", "weak", "unknown"]}, "confidence": {"type": "string", "enum": ["low",
                           "medium", "high"]}, "signalsUsed": ARR, "concerns": ARR}}
    req.append("overallFit")
    return {"$schema": "https://json-schema.org/draft/2020-12/schema", "type": "object",
            "additionalProperties": False, "required": ["companies"],
            "properties": {"companies": {"type": "array", "maxItems": cfg.get("max_per_call", 15),
                           "items": {"type": "object", "additionalProperties": False,
                                     "required": req, "properties": props}}}}


# ----------------------------- queries --------------------------
def _providers(cfg):
    # SKILL.md calls this plan field connect_providers; accept both spellings
    srcs = cfg.get("data_sources") or cfg.get("connect_providers") or []
    return [d.strip() for d in srcs if isinstance(d, str) and d.strip()]


def _hard_list(cfg):
    return "\n".join(f"{i}. {h['text']}" for i, h in enumerate(cfg["hard_criteria"], 1))


def discovery_query(cfg, segment):
    p = [f"Find real companies matching: {cfg['objective']}.",
         "HARD REQUIREMENTS — every company must satisfy ALL of these; do NOT include a company "
         "that fails any of them:\n" + _hard_list(cfg)]
    if cfg.get("soft_criteria"):
        p.append("NICE-TO-HAVE signals (grade these, do not filter on them): "
                 + " ".join(s["text"] for s in cfg["soft_criteria"]))
    p.append("FOCUS for this search (a place to look, NOT ground truth — verify every company "
             f"independently against the hard requirements): {segment['focus']}")
    if cfg.get("time_window"):
        p.append(f"TIME WINDOW: only include companies whose qualifying event happened {cfg['time_window']}. "
                 "Anchor every date claim to a dated source.")
    if cfg.get("seed_company"):
        p.append(f"Do NOT include {cfg['seed_company']} itself — the list is about companies "
                 "around it, not the company.")
    p.append("For EACH company fill every field: identity (official name, canonical website "
             "homepage URL, HQ city/region, a one-line description of what it actually sells — "
             "not its slogan); met yes/no/unknown with concrete public evidence for EACH hard "
             "requirement; a graded level with concrete public signals for EACH nice-to-have; "
             "every requested data column; and an overallFit (tier, confidence, signalsUsed, concerns).")
    if _providers(cfg):
        srcs = ", ".join(_providers(cfg))
        routed = [f"{c['key']} from {c['source']}" for c in cfg.get("columns", [])
                  if c.get("source") not in (None, "", "web")]
        p.append(f"Data sources attached: {srcs}. Use them for the fields that name them"
                 + (f" ({'; '.join(routed)})" if routed else "") + ".")
    p.append("Grade strictly and comparatively so the ranking discriminates: a nice-to-have is "
             "'strong' only with direct public evidence, 'partial'/'medium' when inferred. Reserve "
             "tier 'exceptional' for at most one or two near-perfect fits per batch. Set confidence "
             "'high' only when multiple independent sources corroborate.")
    p.append("Search beyond the company homepage: funding announcements, press coverage, product "
             "docs and pricing pages, GitHub, customer case studies, LinkedIn company pages, job "
             "postings; corroborate across at least two sources where possible.")
    p.append("Use null, empty arrays, or 'unknown' whenever a fact is not supported by public "
             "evidence; NEVER fabricate a company name, website, funding amount, or number. If you "
             "cannot confirm the company's real website, set website to null. Do not include "
             "duplicates, subsidiaries already covered by their parent, or companies that have "
             "shut down or been acquired (unless the objective says otherwise). Do not include any "
             "company from the exclusion list.")
    p.append(f"Return up to {cfg.get('max_per_call', 15)} real companies.")
    return "\n\n".join(p)


def verify_query(cfg):
    checks = ", ".join(cfg.get("verify_columns", [])[:3])
    q = (f"Fact-check this company list for: {cfg['objective']}. The input data has one row per "
         "company with an id and its claimed name, website, HQ, description, and key claimed "
         "facts. For EACH row, use web search"
         + (" and the attached data sources" if _providers(cfg) else "") +
         " to determine: (1) is this a real, currently operating company (not shut down, not "
         "acquired-and-absorbed, not a product name mistaken for a company); (2) does the claimed "
         "website actually belong to this company; (3) for EACH of the following hard "
         "requirements, does the company actually satisfy it TODAY — check each independently "
         "and cite evidence:\n" + _hard_list(cfg))
    if cfg.get("time_window"):
        q += f"\n(The qualifying event must have happened {cfg['time_window']}.)"
    if checks:
        q += f"\n(4) verify or correct the claimed key facts where you can ({checks})."
    keys = ", ".join(f"{i}={h['key']}" for i, h in enumerate(cfg["hard_criteria"], 1))
    q += ("\nBe skeptical: companies drift out of criteria (a 'Series A' company may have raised "
          "a C; an 'independent' company may have been acquired). If you cannot find evidence, "
          "mark it 'unknown' — never guess. Return exactly one verdict per row, copying the "
          "row's id into the verdict's id field unchanged, with one criteria entry per hard "
          f"requirement, setting each entry's key to that requirement's key ({keys}).")
    return q


def verify_schema(cfg):
    """Verdicts echo each input row's id, so results join onto exactly the right company even
    when two companies share a name; each criteria entry echoes its requirement's key (an enum,
    so the join back onto hard criteria is exact — no fuzzy text matching)."""
    return {"$schema": "https://json-schema.org/draft/2020-12/schema", "type": "object",
            "additionalProperties": False, "required": ["verdicts"],
            "properties": {"verdicts": {"type": "array", "items": {
                "type": "object", "additionalProperties": False,
                "required": ["id", "name", "exists", "website_valid", "criteria", "corrections"],
                "properties": {
                    "id": {"type": "string"}, "name": {"type": "string"},
                    "exists": {"type": "string", "enum": ["confirmed", "likely", "uncertain", "not_found"]},
                    "website_valid": {"type": "string", "enum": ["valid", "unverifiable", "wrong"]},
                    "criteria": {"type": "array", "items": {
                        "type": "object", "additionalProperties": False,
                        "required": ["key", "met", "evidence"],
                        "properties": {"key": {"type": "string",
                                               "enum": [h["key"] for h in cfg["hard_criteria"]]},
                                       "met": {"type": "string", "enum": ["yes", "no", "unknown"]},
                                       "evidence": {"type": "string"}}}},
                    "corrections": {"type": "array", "items": {"type": "string"},
                                    "description": "claimed facts that are wrong or stale, with the corrected value and source"}
                }}}}}


# ----------------------------- keys & dedupe --------------------
def norm_domain(url):
    u = (url or "").strip().lower()
    if not u:
        return ""
    u = re.sub(r"^[a-z]+://", "", u)
    u = u.split("/")[0].split("?")[0].split("#")[0].rstrip(".")
    u = re.sub(r"^www\.", "", u)
    return u if "." in u else ""


def norm_name(name):
    n = re.sub(r"[^a-z0-9 ]", "", (name or "").lower()).strip()
    words = n.split()
    while words and words[-1] in LEGAL_SUFFIXES:
        words.pop()
    return " ".join(words)


def _key(c):
    d = norm_domain(c.get("website"))
    return "dom:" + d if d else "nm:" + norm_name(c.get("name"))


def load_existing(cfg, extra_path=None):
    """The user's already-known companies -> (set of dedupe keys, exclusion payload rows).
    Pulls from three places, all optional: the inline `exclude_companies` list, the
    `existing_list.file` named in the config, and an --exclude-file passed on the CLI. Each
    file may be a CSV (uses match_on columns, else any of domain/website/url/company/name) or a
    plain text file of one name-or-domain per line."""
    keys, rows = set(), []
    spec = cfg.get("existing_list") or {}

    def add(name, site):
        d, n = norm_domain(site), norm_name(name)
        if not d and not n:
            return
        if d:
            keys.add("dom:" + d)
        if n:
            keys.add("nm:" + n)
        rows.append({k: v for k, v in (("company", name or d), ("website", d or None)) if v})

    KNOWN_COLS = ("domain", "website", "url", "company", "name")

    def parse_file(path, use_spec_match_on):
        with open(path, newline="", encoding="utf-8-sig") as fh:
            sample = fh.read(4096)
            fh.seek(0)
            lines = sample.splitlines()
            first = lines[0].strip().lower() if lines else ""
            # CSV if the first line has commas OR is a lone known header word (a single-column
            # CSV has no commas but must not be read as a plain name list, header included)
            is_csv = "," in first or first in KNOWN_COLS
            if is_csv:
                reader = csv.DictReader(fh)
                # spec.match_on describes the CONFIG's existing_list file only; a --exclude-file
                # gets its columns derived from its own header so a config match_on can't blank it
                match_on = (spec.get("match_on") if use_spec_match_on else None) \
                    or [c for c in (reader.fieldnames or []) if c.lower() in KNOWN_COLS]
                for r in reader:
                    site = next((r[c] for c in match_on if c in r and "." in (r[c] or "")), None)
                    name = next((r[c] for c in match_on if c in r and r[c] and "." not in r[c]), None)
                    add(name, site)
            else:
                for line in fh.read().splitlines():
                    line = line.strip()
                    if line:
                        add(line if "." not in line else None, line if "." in line else None)

    for item in cfg.get("exclude_companies", []):
        if isinstance(item, str) and item.strip():
            add(item, item if "." in item else None)
    for path, use_spec in ((spec.get("file"), True), (extra_path, False)):
        if not path:
            continue
        if not os.path.exists(path):
            print(f"warning: exclude/existing-list file {path} not found — skipping it")
            continue
        before = len(rows)
        parse_file(path, use_spec)
        if len(rows) == before:
            print(f"warning: {path} yielded no usable company names/domains — check its format "
                  "(one name-or-domain per line, or a CSV with a domain/website/company/name column)")
    return keys, rows


# ----------------------------- orchestration --------------------
def discovery_body(cfg, schema, seg, excl, prev_rid=None):
    if prev_rid:
        q = (f"Find up to {cfg.get('max_per_call', 15)} MORE companies matching the same brief, "
             "excluding every company already returned and everyone in the exclusion list. Apply "
             "the same hard requirements, grading, and anti-fabrication rules as before.")
    else:
        q = discovery_query(cfg, seg)
    body = {"query": q, "effort": cfg.get("discovery_effort", "medium"), "outputSchema": schema}
    if prev_rid:
        body["previousRunId"] = prev_rid
    if excl:
        body["input"] = {"exclusion": excl}
    if _providers(cfg):
        body["dataSources"] = [{"provider": p} for p in _providers(cfg)]
    return body


def discover(cfg, schema, segments, concurrency, existing_rows, prev_run_ids=None, seen=None):
    """One run per segment. Returns (companies, {segment label: run id}); the run ids feed
    --more / --rounds continuations via previousRunId."""
    found = []
    seen = list(seen or [])          # [{"company":..., "website":...}] of already-found rows
    run_ids = {}
    for i in range(0, len(segments), concurrency):
        batch = segments[i:i + concurrency]
        # cap keeps create bodies bounded, but the existing list must never be pushed out of
        # the window by this session's finds: it gets priority, newest finds fill what's left
        # (at least half the cap once the pool grows), and local dedupe catches the overflow
        cap = cfg.get("exclusion_cap", 100)
        keep_seen = min(len(seen), cap // 2 if existing_rows else cap)
        excl = existing_rows[:cap - keep_seen] + (seen[-keep_seen:] if keep_seen else [])
        started = []
        for seg in batch:   # create sequentially: a parallel burst can trip the account QPS limit
            rid = create_run(discovery_body(cfg, schema, seg, excl, (prev_run_ids or {}).get(seg["label"])),
                             label=seg["label"])
            if rid:
                started.append((seg, rid))
                run_ids[seg["label"]] = rid
            time.sleep(1)
        results = {}

        def work(seg, rid):
            r = wait_run(rid, label=seg["label"])
            out = ((r or {}).get("output") or {})
            cos = (out.get("structured") or {}).get("companies") or []
            grounded = grounding_by_index(out.get("grounding"))
            for j, c in enumerate(cos):
                c["_segment"] = seg["label"]
                if j in grounded:
                    c["_sources"] = grounded[j]
            results[seg["label"]] = cos
            print(f"  [{seg['label']}] {len(cos)} companies")

        ts = [threading.Thread(target=work, args=(s, rid)) for s, rid in started]
        for t in ts: t.start()
        for t in ts: t.join()
        for s, _rid in started:
            for c in results.get(s["label"], []):
                found.append(c)
                row = {k: v for k, v in (("company", c.get("name")),
                                         ("website", norm_domain(c.get("website")) or None)) if v}
                if row:
                    seen.append(row)
    return found, run_ids


_CO_FIELD = re.compile(r"^structured\.companies\[(\d+)\](?:\.|$)")


def grounding_by_index(grounding):
    """Attribute a run's output.grounding citations to company indexes. Field paths arrive at
    three granularities: "structured.companies[3]" (whole row), "structured.companies[3].name"
    (one field, rolled up to its row), and bare "structured" (run-level research trail, dropped:
    it supports no one company's claims). Citation URLs dedup per company. Returns {index: urls}."""
    by = {}
    if not isinstance(grounding, list):
        return by
    for e in grounding:
        if not isinstance(e, dict):
            continue
        m = _CO_FIELD.match(e.get("field") or "")
        cites = e.get("citations")
        if not m or not isinstance(cites, list):
            continue
        i = int(m.group(1))
        urls = by.setdefault(i, [])
        for c in cites:
            u = c.get("url") if isinstance(c, dict) else None
            if isinstance(u, str) and u.strip() and u not in urls:
                urls.append(u)
    return {i: urls for i, urls in by.items() if urls}


def merge_dupes(cfg, keep, other):
    """Two rows for the same company: `keep` is the higher-scored copy; fill its null/empty
    identity and column fields from the other copy and keep both copies' citations."""
    for k in ["website", "hq", "description"] + [c["key"] for c in cfg.get("columns", [])]:
        if keep.get(k) in (None, "", []) and other.get(k) not in (None, "", []):
            keep[k] = other[k]
    keep["_sources"] = list(dict.fromkeys((keep.get("_sources") or []) + (other.get("_sources") or [])))
    return keep


def met(c, k):
    return ((c.get(k) or {}).get("met")) or "unknown"


def lvl(c, k):
    return ((c.get(k) or {}).get("level")) or "unknown"


def score(cfg, c):
    capw = {"strong": 2, "partial": 1, "none": 0, "unknown": 0}
    strw = {"strong": 2, "medium": 1.5, "weak": 1, "none": 0, "unknown": 0}
    s = 0
    for d in cfg.get("soft_criteria", []):
        w = capw if d.get("scale", "strength") == "capability" else strw
        s += w.get(lvl(c, d["key"]), 0)
    tier = {"exceptional": 90, "strong": 76, "moderate": 58, "weak": 38, "unknown": 50}.get(
        (c.get("overallFit") or {}).get("tier", "unknown"), 50)
    conf = {"high": 6, "medium": 0, "low": -6}.get((c.get("overallFit") or {}).get("confidence", "medium"), 0)
    hq = (c.get("hq") or "").lower()
    pref = (cfg.get("geography") or {}).get("preferred_regions") or []
    # full first-segment containment, not a truncated prefix ("united" would match both
    # United States and United Kingdom)
    in_pref = any(p.lower().split(",")[0].strip() in hq for p in pref) if hq and pref else False
    return tier + conf + s * 1.4 + (4 if in_pref else 0), in_pref


def max_possible_score(cfg):
    """Displayed scores are a percentage of the best score the rubric allows, not a clip at 100:
    clipping collapses every good company onto the same number."""
    return 90 + 6 + len(cfg.get("soft_criteria", [])) * 2 * 1.4 \
        + (4 if cfg.get("geography", {}).get("preferred_regions") else 0)


def calibrate(cfg, c):
    """Discount thin/unverified rows so they don't top the list."""
    base, in_pref = c["_score"], c["_pref"]
    pen = 0
    if not norm_domain(c.get("website")): pen += 15
    if not (c.get("hq") or "").strip() and (cfg.get("geography") or {}): pen += 10
    if (c.get("overallFit") or {}).get("confidence") == "low": pen += 10
    # an unassessable fit shouldn't outrank an honestly-graded weak one (unknown tier
    # scores 50 base vs weak's 38, so discount it here)
    if (c.get("overallFit") or {}).get("tier") == "unknown": pen += 8
    unknown_hard = sum(1 for h in cfg["hard_criteria"] if met(c, h["key"]) == "unknown")
    pen += min(unknown_hard * 5, 10)
    if c.get("_exists") == "not_found": pen += 25
    elif c.get("_exists") == "uncertain": pen += 8
    if c.get("_vsite") == "wrong": pen += 10
    if any(v == "unknown" for v in (c.get("_vcriteria") or {}).values()): pen += 8
    cal = (base - pen) * 100 / max_possible_score(cfg)
    if not norm_domain(c.get("website")):
        cal = min(cal, 82)
    return round(min(100, max(0, cal)))


def verify(cfg, companies, concurrency, verdicts=None, checkpoint=None):
    """Fact-check in batches of 8. Each shortlist row travels as an input.data row carrying the
    company's stable key as id; the verdict echoes it back, so the join is exact even when two
    companies share a name. Updates and returns {company key: verdict}; calls checkpoint()
    after each batch group so paid-for verdicts survive a crash."""
    batches = [companies[i:i + 8] for i in range(0, len(companies), 8)]
    schema = verify_schema(cfg)
    verdicts, lock = verdicts if verdicts is not None else {}, threading.Lock()

    def claimed_facts(c):
        bits = []
        for name in cfg.get("verify_columns", [])[:3]:
            v = c.get(name)
            if v not in (None, "", []):
                bits.append(f"{name}: {v}")
        return "; ".join(bits)

    for i in range(0, len(batches), concurrency):
        grp = batches[i:i + concurrency]
        started = []
        for b in grp:
            rows = [{"id": _key(c), "name": c.get("name"), "website": c.get("website"),
                     "claimed_hq": c.get("hq"), "claimed_description": c.get("description"),
                     "claimed_facts": claimed_facts(c)} for c in b]
            body = {"query": verify_query(cfg), "effort": cfg.get("verify_effort", "high"),
                    "outputSchema": schema, "input": {"data": rows}}
            if _providers(cfg):
                body["dataSources"] = [{"provider": p} for p in _providers(cfg)]
            rid = create_run(body, label="verify")
            if rid:
                started.append((b, rid))
            time.sleep(1)

        def work(b, rid):
            r = wait_run(rid, label="verify")
            sent = {_key(c) for c in b}
            for v in (((r or {}).get("output") or {}).get("structured") or {}).get("verdicts") or []:
                if v.get("id") in sent:
                    with lock:
                        verdicts[v["id"]] = v
                else:
                    print(f"  dropped verdict with unknown id {v.get('id')!r} ({v.get('name')})")

        ts = [threading.Thread(target=work, args=(b, rid)) for b, rid in started]
        for t in ts: t.start()
        for t in ts: t.join()
        print(f"  verified {min((i + concurrency) * 8, len(companies))}/{len(companies)}")
        if checkpoint:
            checkpoint()
    return verdicts


def join_criteria(cfg, verdict):
    """Map a verdict's criteria entries back to hard-criterion keys. The schema forces each
    entry's key to one of the config keys (enum), so this is an exact join; order is a
    fallback for a malformed reply. Returns {key: met}."""
    out = {}
    entries = [e for e in (verdict.get("criteria") or []) if isinstance(e, dict)]
    hards = cfg["hard_criteria"]
    for e in entries:
        if e.get("key") and e["key"] not in out:
            out[e["key"]] = e.get("met", "unknown")
    if len(out) < len(hards) and len(entries) == len(hards):   # order fallback
        for h, e in zip(hards, entries):
            out.setdefault(h["key"], e.get("met", "unknown"))
    for h in hards:
        out.setdefault(h["key"], "unknown")
    return out


# ----------------------------- output ---------------------------
def _cell(v):
    if v is None:
        return ""
    if isinstance(v, bool):
        return "true" if v else "false"
    if isinstance(v, list):
        return " | ".join(str(x) for x in v)
    return v


def write_outputs(cfg, final, table_cols):
    col_keys = [c["key"] for c in cfg.get("columns", [])]
    soft_keys = [s["key"] for s in cfg.get("soft_criteria", [])]
    hard_keys = [h["key"] for h in cfg["hard_criteria"]]
    cols = ["rank", "company", "website", "hq", "description", "score"] + col_keys \
        + soft_keys + hard_keys + ["overall_tier", "confidence", "concerns",
                                   "verify_exists", "verify_website", "corrections", "sources", "segment"]

    def row(i, c):
        of = c.get("overallFit") or {}
        vc = c.get("_vcriteria") or {}
        return [i, c.get("name"), c.get("website"), c.get("hq"), c.get("description"), c["_calib"]] \
            + [_cell(c.get(k)) for k in col_keys] \
            + [lvl(c, k) for k in soft_keys] \
            + [vc.get(k) or met(c, k) for k in hard_keys] \
            + [of.get("tier"), of.get("confidence"), " | ".join((of.get("concerns") or [])[:2]),
               c.get("_exists", ""), c.get("_vsite", ""),
               " | ".join(c.get("_corrections") or []),
               " | ".join(c.get("_sources") or []), c.get("_segment")]

    with open("companies.csv", "w", newline="") as fh:
        w = csv.writer(fh); w.writerow(cols)
        for i, c in enumerate(final, 1):
            w.writerow(row(i, c))
    print(f"wrote companies.csv ({len(final)} rows)")
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Companies"
        ws.append(cols)
        for c in ws[1]:
            c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor="1F3864")
        for i, c in enumerate(final, 1):
            ws.append(row(i, c))
        ws.freeze_panes = "C2"
        wb.save("companies.xlsx"); print("wrote companies.xlsx")
    except ImportError:
        print("(openpyxl not installed — skipped .xlsx; `pip install openpyxl` for the formatted sheet)")
    try:
        try:
            import render_viewer
        except ImportError:
            sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
            import render_viewer
        render_viewer.render("companies.csv", "companies.html",
                             title=cfg.get("objective") or "Companies", table_cols=table_cols)
    except (Exception, SystemExit) as e:
        print(f"(companies.html skipped: {e})")


# ----------------------------- main -----------------------------
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--config", required=True)
    ap.add_argument("--target", type=int, default=0, help="how many companies to keep (default: config target_count, else 25)")
    ap.add_argument("--no-verify", action="store_true")
    ap.add_argument("--limit-segments", type=int, default=0, help="use only the first N segments (smoke test)")
    ap.add_argument("--max-concurrent", type=int, default=0, help="override config concurrency")
    ap.add_argument("--rounds", type=int, default=1,
                    help="discovery rounds per segment; >1 keeps digging via previousRunId until dry "
                         "(for exhaustive 'find EVERY company' asks)")
    ap.add_argument("--more", action="store_true",
                    help="continue the previous session's runs (previousRunId) and fetch new companies")
    ap.add_argument("--state", default="research_state.json", help="session state file, enables --more")
    ap.add_argument("--exclude-file",
                    help="file of company names or domains (one per line, or a CSV) to skip — an "
                         "existing list; merged with the config's exclude_companies / existing_list")
    a = ap.parse_args()
    cfg = json.load(open(a.config))
    validate_config(cfg)
    target = a.target or cfg.get("target_count", 25)
    exhaustive = cfg.get("mode") == "exhaustive"
    if exhaustive and a.rounds == 1:
        print("note: mode=exhaustive with --rounds 1 — pass --rounds 3+ to keep digging until segments run dry")
    segments = cfg["segments"][:a.limit_segments] if a.limit_segments else cfg["segments"]
    conc = a.max_concurrent or cfg.get("concurrency", 2)
    schema = build_schema(cfg)
    existing_keys, existing_rows = load_existing(cfg, a.exclude_file)
    if existing_keys:
        print(f"== existing list: {len(existing_rows)} companies to dedupe against ==")

    state = {"run_ids": {}, "pool": [], "verdicts": {}}
    if a.more:
        if not os.path.exists(a.state):
            sys.exit(f"ERROR: --more continues a previous session, but {a.state} was not found.")
        state = json.load(open(a.state))

    by = {_key(c): c for c in state["pool"]}
    seen = [{k: v for k, v in (("company", c.get("name")),
                               ("website", norm_domain(c.get("website")) or None)) if v}
            for c in by.values()]

    run_ids = dict(state.get("run_ids", {}))

    def save_state():
        # checkpoint so a crash mid-verify doesn't lose paid-for discovery/verdicts;
        # --more resumes from whatever was saved last
        json.dump({"run_ids": run_ids, "pool": sorted(by.values(), key=lambda x: -x["_score"]),
                   "verdicts": verdicts}, open(a.state, "w"))

    verdicts = state.get("verdicts", {})
    raw_total, dropped_existing, dry_rounds = 0, 0, 0
    for rnd in range(1, max(1, a.rounds) + 1):
        cont = a.more or rnd > 1
        print(f"== discovery round {rnd}/{a.rounds}: {len(segments)} segments, {conc} concurrent"
              f"{' (continuation)' if cont else ''} ==")
        raw, rids = discover(cfg, schema, segments, conc, existing_rows,
                             prev_run_ids=run_ids if cont else None, seen=seen)
        run_ids.update(rids)
        raw_total += len(raw)
        new = 0
        for c in raw:
            if any(met(c, h["key"]) == "no" for h in cfg["hard_criteria"]):
                continue    # discovery was told not to return these; drop stragglers
            k = _key(c)
            if k in existing_keys or ("nm:" + norm_name(c.get("name"))) in existing_keys:
                dropped_existing += 1
                continue
            c["_score"], c["_pref"] = score(cfg, c)
            if k.replace("nm:", "").strip():
                if k not in by:
                    new += 1
                    by[k] = c
                elif c["_score"] > by[k]["_score"]:
                    by[k] = merge_dupes(cfg, c, by[k])
                else:
                    by[k] = merge_dupes(cfg, by[k], c)
                row = {kk: vv for kk, vv in (("company", c.get("name")),
                                             ("website", norm_domain(c.get("website")) or None)) if vv}
                if row:
                    seen.append(row)
        print(f"== round {rnd}: +{new} new unique ==")
        save_state()
        dry_rounds = dry_rounds + 1 if new < 3 else 0
        if rnd < a.rounds and dry_rounds >= 2:
            print("== two consecutive rounds under 3 new companies; stopping early (segments look dry) ==")
            break
    pool = sorted(by.values(), key=lambda x: -x["_score"])
    print(f"== {raw_total} raw -> {len(pool)} unique ({dropped_existing} already on the existing list) ==")

    shortlist = pool if exhaustive else pool[:max(target + 10, target)]
    if not a.no_verify and shortlist:
        unchecked = [c for c in shortlist if _key(c) not in verdicts]
        if unchecked:
            print(f"== verification: {len(unchecked)} of top {len(shortlist)} (effort {cfg.get('verify_effort','high')}) ==")
            verify(cfg, unchecked, conc, verdicts, checkpoint=save_state)
    for c in shortlist:
        v = verdicts.get(_key(c), {})
        c["_exists"] = v.get("exists", "unchecked")
        c["_vsite"] = v.get("website_valid", "unchecked")
        c["_vcriteria"] = join_criteria(cfg, v) if v else {}
        c["_corrections"] = v.get("corrections") or []
    erank = {"confirmed": 0, "likely": 1, "uncertain": 2, "unchecked": 2, "not_found": 3}

    def crank(c):
        vals = [(c.get("_vcriteria") or {}).get(h["key"]) or met(c, h["key"]) for h in cfg["hard_criteria"]]
        if any(v == "no" for v in vals):
            return 3
        return 0 if all(v == "yes" for v in vals) else 1

    def eligible(c):
        if c.get("_exists", "unchecked") == "not_found":
            return False
        if any(v == "no" for v in (c.get("_vcriteria") or {}).values()):
            return False
        if not norm_domain(c.get("website")) and c.get("_exists") != "confirmed":
            return False
        return True

    near_misses = [c for c in shortlist
                   if any(v == "no" for v in (c.get("_vcriteria") or {}).values())]
    for c in near_misses:
        failed = [h["key"] for h in cfg["hard_criteria"] if (c.get("_vcriteria") or {}).get(h["key"]) == "no"]
        print(f"  near miss: {c.get('name')} — failed {', '.join(failed)}")

    for c in shortlist:
        c["_calib"] = calibrate(cfg, c)
    elig = sorted([c for c in shortlist if eligible(c)],
                  key=lambda c: (erank.get(c.get("_exists", "unchecked"), 2), crank(c), -c["_calib"]))
    # exhaustive asks deliver completeness, not a top-N: only trim when a target was set explicitly
    final = elig if exhaustive and not a.target else elig[:target]
    print(f"== final: {len(final)} companies ==")
    table_cols = cfg.get("table_cols") or [c["key"] for c in cfg.get("columns", [])[:3]]
    write_outputs(cfg, final, table_cols)

    save_state()
    print(f"wrote {a.state} (rerun with --more to fetch additional companies)")


if __name__ == "__main__":
    main()
